"""
Geração do relatório Excel com 3 abas.
Adaptado de gerar_relatorio.py com empresa_name dinâmico.
"""
from datetime import datetime
from pathlib import Path


def calcular_stats(reviews: list[dict]) -> dict:
    notas = [r["nota"] for r in reviews if isinstance(r["nota"], int) and r["nota"] > 0]
    media = sum(notas) / len(notas) if notas else 0
    dist = {i: notas.count(i) for i in range(1, 6)}
    com_texto = sum(1 for r in reviews if r.get("review", "").strip())
    com_resposta = sum(1 for r in reviews if r.get("resposta_dono", "").strip())
    return {
        "total": len(reviews),
        "media": media,
        "dist": dist,
        "notas": notas,
        "com_texto": com_texto,
        "com_resposta": com_resposta,
    }


def gerar_excel(reviews: list[dict], empresa_name: str, output_dir: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    import re as _re
    stats = calcular_stats(reviews)
    ts = datetime.now().strftime("%Y%m%d")

    # Sanitiza o nome para uso no sistema de arquivos (remove caracteres inválidos no Windows)
    safe_name = _re.sub(r'[\\/:*?"<>|]', "_", empresa_name).strip()
    safe_name = safe_name[:60]  # limite de comprimento

    empresa_dir = output_dir / safe_name
    empresa_dir.mkdir(parents=True, exist_ok=True)

    COR_HEADER    = "1F4E79"
    COR_SUBHEADER = "2E75B6"
    COR_5STAR     = "375623"
    COR_4STAR     = "538135"
    COR_3STAR     = "7F6000"
    COR_2STAR     = "C55A11"
    COR_1STAR     = "C00000"
    COR_ZEBRA     = "EBF3FB"
    COR_NEG_BG    = "FCE4D6"
    STAR_COLORS   = {5: COR_5STAR, 4: COR_4STAR, 3: COR_3STAR, 2: COR_2STAR, 1: COR_1STAR}

    def header_style(cell, bg=COR_HEADER, size=11, bold=True):
        cell.font = Font(bold=bold, color="FFFFFF", size=size, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def border_all(cell):
        thin = Side(style="thin", color="BFBFBF")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    # ── ABA 1: RESUMO ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14

    ws.merge_cells("A1:C1")
    ws["A1"] = "Relatório de Avaliações Google Maps"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    ws["A1"].fill = PatternFill("solid", fgColor=COR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    ws["A2"] = f"{empresa_name}  •  Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="595959", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 8

    metricas = [
        ("Total de avaliações", stats["total"]),
        ("Média geral", f"{stats['media']:.2f} / 5.00"),
        ("Com comentário", stats["com_texto"]),
        ("Sem comentário", stats["total"] - stats["com_texto"]),
        ("Com resposta do proprietário", stats["com_resposta"]),
    ]
    ws["A4"] = "Métrica"
    ws["B4"] = "Valor"
    header_style(ws["A4"], COR_SUBHEADER, size=10)
    header_style(ws["B4"], COR_SUBHEADER, size=10)
    border_all(ws["A4"]); border_all(ws["B4"])
    ws.row_dimensions[4].height = 20

    for i, (label, valor) in enumerate(metricas, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = valor
        ws[f"A{i}"].font = Font(name="Calibri", size=10)
        ws[f"B{i}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{i}"].alignment = Alignment(horizontal="center")
        if i % 2 == 0:
            ws[f"A{i}"].fill = PatternFill("solid", fgColor=COR_ZEBRA)
            ws[f"B{i}"].fill = PatternFill("solid", fgColor=COR_ZEBRA)
        border_all(ws[f"A{i}"]); border_all(ws[f"B{i}"])
    ws.row_dimensions[9].height = 4

    ws["A11"] = "Estrelas"
    ws["B11"] = "Quantidade"
    ws["C11"] = "Percentual"
    header_style(ws["A11"], COR_SUBHEADER, size=10)
    header_style(ws["B11"], COR_SUBHEADER, size=10)
    header_style(ws["C11"], COR_SUBHEADER, size=10)
    border_all(ws["A11"]); border_all(ws["B11"]); border_all(ws["C11"])
    ws.row_dimensions[11].height = 20

    notas = stats["notas"]
    for i, s in enumerate([5, 4, 3, 2, 1], start=12):
        qtd = stats["dist"][s]
        pct = qtd / len(notas) * 100 if notas else 0
        ws[f"A{i}"] = f"{'★' * s} ({s} estrelas)"
        ws[f"B{i}"] = qtd
        ws[f"C{i}"] = f"{pct:.1f}%"
        ws[f"A{i}"].font = Font(name="Calibri", size=10, color=STAR_COLORS[s], bold=True)
        ws[f"B{i}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"B{i}"].alignment = Alignment(horizontal="center")
        ws[f"C{i}"].font = Font(name="Calibri", size=10)
        ws[f"C{i}"].alignment = Alignment(horizontal="center")
        for col in ["A", "B", "C"]:
            border_all(ws[f"{col}{i}"])
        ws.row_dimensions[i].height = 18

    ws["E4"] = "Tipo"
    ws["F4"] = "Qtd"
    header_style(ws["E4"], COR_SUBHEADER, size=10)
    header_style(ws["F4"], COR_SUBHEADER, size=10)
    border_all(ws["E4"]); border_all(ws["F4"])
    ws["E5"] = "Com comentário"
    ws["F5"] = stats["com_texto"]
    ws["E6"] = "Sem comentário"
    ws["F6"] = stats["total"] - stats["com_texto"]
    for row in [5, 6]:
        ws[f"E{row}"].font = Font(name="Calibri", size=10)
        ws[f"F{row}"].font = Font(name="Calibri", size=10, bold=True)
        ws[f"F{row}"].alignment = Alignment(horizontal="center")
        border_all(ws[f"E{row}"]); border_all(ws[f"F{row}"])

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Distribuição por Estrelas"
    chart1.y_axis.title = "Quantidade"
    chart1.style = 10
    chart1.width = 14
    chart1.height = 10
    chart1.grouping = "clustered"
    data1 = Reference(ws, min_col=2, min_row=11, max_row=16)
    cats1 = Reference(ws, min_col=1, min_row=12, max_row=16)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.series[0].graphicalProperties.solidFill = "2E75B6"
    ws.add_chart(chart1, "E8")

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Com vs Sem Comentário"
    chart2.style = 10
    chart2.width = 10
    chart2.height = 7
    data2 = Reference(ws, min_col=6, min_row=4, max_row=6)
    cats2 = Reference(ws, min_col=5, min_row=5, max_row=6)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    chart2.series[0].graphicalProperties.solidFill = "375623"
    ws.add_chart(chart2, "E26")

    # ── ABA 2: TODAS AS AVALIAÇÕES ─────────────────────────────────────────
    ws2 = wb.create_sheet("Avaliações")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 60
    ws2.column_dimensions["F"].width = 50

    headers = ["#", "Nome", "Nota", "Data", "Avaliação", "Resposta do Proprietário"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        header_style(cell, COR_HEADER)
        border_all(cell)
    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = "A2"

    for i, r in enumerate(reviews, start=1):
        row = i + 1
        nota = r["nota"] if isinstance(r["nota"], int) else 0
        estrelas_txt = "★" * nota + "☆" * (5 - nota)

        ws2.cell(row=row, column=1, value=i)
        ws2.cell(row=row, column=2, value=r.get("nome", ""))
        ws2.cell(row=row, column=3, value=estrelas_txt)
        ws2.cell(row=row, column=4, value=r.get("data", ""))
        ws2.cell(row=row, column=5, value=r.get("review", ""))
        ws2.cell(row=row, column=6, value=r.get("resposta_dono", ""))

        bg = COR_NEG_BG if nota <= 2 else ("FFFFFF" if i % 2 == 0 else COR_ZEBRA)
        star_color = STAR_COLORS.get(nota, "000000")

        for col in range(1, 7):
            cell = ws2.cell(row=row, column=col)
            cell.font = Font(
                name="Calibri", size=9,
                color=star_color if col == 3 else "000000",
                bold=(col == 3),
            )
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="top", wrap_text=(col >= 5))
            border_all(cell)

        ws2.row_dimensions[row].height = 40 if r.get("review") else 18

    # ── ABA 3: NEGATIVOS ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Negativos (1-2★)")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 65
    ws3.column_dimensions["F"].width = 55

    for col, h in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=h)
        header_style(cell, COR_1STAR)
        border_all(cell)
    ws3.row_dimensions[1].height = 22
    ws3.freeze_panes = "A2"

    negativos = [r for r in reviews if isinstance(r["nota"], int) and r["nota"] <= 2]
    for i, r in enumerate(negativos, start=1):
        row = i + 1
        nota = r["nota"]
        estrelas_txt = "★" * nota + "☆" * (5 - nota)

        ws3.cell(row=row, column=1, value=i)
        ws3.cell(row=row, column=2, value=r.get("nome", ""))
        ws3.cell(row=row, column=3, value=estrelas_txt)
        ws3.cell(row=row, column=4, value=r.get("data", ""))
        ws3.cell(row=row, column=5, value=r.get("review", ""))
        ws3.cell(row=row, column=6, value=r.get("resposta_dono", ""))

        for col in range(1, 7):
            cell = ws3.cell(row=row, column=col)
            cell.font = Font(
                name="Calibri", size=9,
                color=STAR_COLORS.get(nota, "000000") if col == 3 else "000000",
                bold=(col == 3),
            )
            cell.fill = PatternFill("solid", fgColor=COR_NEG_BG if i % 2 == 0 else "FFF2CC")
            cell.alignment = Alignment(vertical="top", wrap_text=(col >= 5))
            border_all(cell)
        ws3.row_dimensions[row].height = 55

    xlsx_path = empresa_dir / f"{safe_name}_Relatorio_Avaliacoes_{ts}.xlsx"
    wb.save(str(xlsx_path))
    return xlsx_path
